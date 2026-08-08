import openpyxl 
from openpyxl.styles import Font, PatternFill, Alignment 
from datetime import datetime 
 
def generate_beneficiary_excel(beneficiaries, filename=None): 
    if filename is None: 
        filename = f"beneficiary_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx" 
 
    wb = openpyxl.Workbook() 
    ws = wb.active 
    ws.title = "Beneficiaries" 
 
    # Headers 
    headers = ["ID", "Name", "DOB", "Gender", "Contact", "Email", "Village", "Block", "District", "Education", "Occupation", "Income"] 
    for col, header in enumerate(headers, 1): 
        cell = ws.cell(row=1, column=col, value=header) 
        cell.font = Font(bold=True) 
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid") 
 
    # Data 
    for row, b in enumerate(beneficiaries, 2): 
        ws.cell(row=row, column=1, value=b.get('beneficiary_id', '')) 
        ws.cell(row=row, column=2, value=b.get('full_name', '')) 
        ws.cell(row=row, column=3, value=str(b.get('date_of_birth', ''))) 
        ws.cell(row=row, column=4, value=b.get('gender', '')) 
        ws.cell(row=row, column=5, value=b.get('contact_number', '')) 
        ws.cell(row=row, column=6, value=b.get('email', '')) 
        ws.cell(row=row, column=7, value=b.get('village', '')) 
        ws.cell(row=row, column=8, value=b.get('block', '')) 
        ws.cell(row=row, column=9, value=b.get('district', '')) 
        ws.cell(row=row, column=10, value=b.get('education_level', '')) 
        ws.cell(row=row, column=11, value=b.get('occupation', '')) 
        ws.cell(row=row, column=12, value=float(b.get('family_income', 0))) 
 
    # Auto-adjust column widths 
    for col in range(1, 13): 
        ws.column_dimensions[chr(64 + col)].width = 18 
 
    wb.save(filename) 
    return filename 
